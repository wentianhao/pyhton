"""
时间间隔十分钟，8月15号8点到10点的数据,存于data4.xlsx中
data2.xlsx中存的是数据
取不同的数据更改sheetname即可，存于不同的文件也需改sheetname
"""


import xlrd,xlwt
import datetime
import random
import numpy as np
import pandas as pd
from openpyxl import load_workbook

class getuser():
    def getusers(self):
        # 打开需要操作的excel,由于地图左下角区域没有用户，所以不处理那些区域（原点定为（8000，7000））
        excel = xlrd.open_workbook("data2.xlsx")

        # 获取excel的sheet表
        sheet = excel.sheet_by_name("819")
        starttime = datetime.datetime(2016,8,19,7,00,00)
        endtime = datetime.datetime(2016,8,19,7,10,00)


        # 统计截至日期
        deadday = datetime.datetime(2016, 8,19, 20, 00, 00)
        onetime = []
        one = []
        user = []
        print("-------start----")
        print("start", starttime)
        print("end", endtime)

        for i in range(1,sheet.nrows):
            row_i = sheet.row_values(i)
            d = xlrd.xldate_as_datetime(row_i[0],0)
            print(d,i,": ",row_i)
            onetime = []
            # count=0
            # while (count==0):
            if d >= starttime and d < endtime:
                # count+=1
                onetime.append(round(row_i[2],0))
                onetime.append(round(row_i[4],0))
                onetime.append(round(row_i[7],0))
                onetime.append(round(row_i[9],0))
                onetime.append(np.random.normal(0,1))
                # onetime.append(10000)
                onetime.append(-1)
                onetime.append(-1)
                one.append(onetime)
                # print("d",d,"count",count)
                print("onetime:",onetime)
            else:
                # print("one:",one)
                print("len(one)",len(one))
                # if len(one) != 0:
                user.append(one)
                # print("user:",user)
                one = []
                print("                d",d)
                if d >= endtime:
                    starttime = endtime
                    print("start", starttime)
                    endtime = endtime + datetime.timedelta(minutes=10)
                    print("end", endtime)


            if starttime >= deadday:
                d = xlrd.xldate_as_datetime(row_i[0], 0)
                # print("d:", d)
                break
            if i==sheet.nrows-1:
                user.append(one)
        init_user = np.array(user)
        # data = pd.DataFrame(init_user)
        # # 写入excel文件
        # writer = pd.ExcelWriter("./user.xlsx", engine="openpyxl")
        # book = load_workbook("./user.xlsx")
        # writer.book = book
        # data.to_excel(writer, sheet_name='1081720', float_format='%.5f')
        # writer.save()
        # writer.close()



        temp=0
        print("len(user",len(user))
        for i in range (len(user)):
            temp+=len(user[i])
            print("每个时间段的user的个数",i,len(user[i]))
        print("len(user)",temp)

        return user

    def usertoregion(self,user,region,cell,celllength):
        b=0
        for i in range(len(user)):
            if (user[i][0] == cell * celllength and user[i][1] == cell * celllength):
                a = int(cell*cell - 1)
            elif (user[i][0] == cell * celllength):
                a = int(user[i][1] / celllength) * cell + int(user[i][0] / celllength) - 1
            elif (user[i][1] == cell * celllength):
                a = int(user[i][1] / celllength) * cell + int(user[i][0] / celllength) - cell
            else:
                a = int(user[i][1] / celllength) * cell + int(user[i][0] / celllength)
            # print(a)
            if (a < cell*cell):
                region[a] += 1

            if(b<a):
                b=a
        # print(b)
        return region

def getregion():
    users=getuser().getusers()
    print("len(user)",len(users))
    cell=4
    celllength=750
    T=len(users)

    init_region=[[0 for i in range (cell*cell)]for t in range(T)]
    # print("initregion",init_region)
    for t in range(T):
        # print("user[t]",len(users[t]),users[t])
        init_region[t]=getuser().usertoregion(users[t],init_region[t],cell,celllength)
        # print(init_region)
    init_region15 = np.array(init_region)
    data = pd.DataFrame(init_region15)
    # 写入excel文件
    writer = pd.ExcelWriter("./userregion.xlsx",engine="openpyxl")
    book=load_workbook("./userregion.xlsx")
    writer.book=book
    data.to_excel(writer, sheet_name='10819720', float_format='%.5f')
    writer.save()
    writer.close()


    # print(region)
getregion()




